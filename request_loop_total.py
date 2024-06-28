import requests
import json
import time
from openpyxl import Workbook
from datetime import datetime

# 获取当前时间字符串
time_str = datetime.now().strftime('%Y%m%d%H%M%S')

# 循环发送请求
def send_requests_and_save_responses_loop(url, data, headers, num_requests):
    # 响应结果拼接数组
    concat_response_list = []
    # 循环发送POST请求
    for i in range(num_requests):
        try:
            cur_page = str(i+1)
            data = '{"pageSize": 20,"curPage": '+ cur_page +'}'
            # 发送POST请求
            response = requests.post(url, data=data, headers=headers)
            # 获取响应结果
            response_text = response.text
            # 解析JSON数据
            parsed_data = json.loads(response_text)
            if parsed_data['errorCode'] == 0:
                # 追加数组
                return_value = parsed_data['returnValue']
                print(f"请求 {i + 1} 次成功，当前页：{return_value['curPage']}，从{return_value['fromIndex']}行开始 ")
                concat_response_list.extend(return_value['pageList'])
            else:
                print(f"请求 {i + 1} 次失败，原因为： {parsed_data['errorMsg']} ")
        except Exception as e:
            print(f"请求 {i + 1} 次异常，原因为: {e}")
        # 请求等待1s
        time.sleep(1)
    print(f"加载到数据的行数: {len(concat_response_list)}")
    return concat_response_list

# 响应内容写入到本地文本
def response_write_file(concat_response_list):
    if len(concat_response_list) > 0:
        concat_response_str = json.dumps(concat_response_list, ensure_ascii=False)
        filename = f"response_result_{time_str}.json"
        with open(filename, "w", encoding="utf-8") as file:
            file.write(concat_response_str)
        print(f"合并结果已经写入到文件：{filename}")


# 响应数据格式化
def response_data_format(concat_response_list):
    # 数据整合
    result_college_list = []
    for index, each_response in enumerate(concat_response_list):
        # 省份编码
        # provinceCode = each_response['provinceCode']
        # 学校编号
        # collegeBJCode = each_response['collegeBJCode']
        # 专业组编号
        # majorGroupCode = each_response['majorGroupCode']

        # 省份名称
        provinceName = each_response['provinceName']
        # 学校名称
        collegeName = each_response['collegeName']
        # 专业组名称
        majorGroupName = each_response['majorGroupName']
        # 最低录取分数线
        minAdmissionScoreLine = each_response['minAdmissionScoreLine']
        # 最低录取排名
        minAdmissionRank = each_response['minAdmissionRank']
        # 总计划录取数-各专业计划录取数之和
        totalPlanCount = 0
        # 总计划录制数明细
        totalPlanCountMsg = ""
        # 包含专业
        stuCountRanges = each_response['stuCountRanges']
        # 包含的专业个数
        majorCount = len(stuCountRanges)
        for each_stu in stuCountRanges:
            # 专业名称
            majorName = each_stu['majorName']
            # 计划人数
            planCount = each_stu['planCount']
            # 计划人数累加
            totalPlanCount += planCount
            # 专业计划人数详细拼接
            totalPlanCountMsg += majorName + ':' + str(planCount) + ";"
        # 顺序放入数组
        row = [index + 1, provinceName, collegeName, majorGroupName, "共" + str(majorCount) +"个专业", totalPlanCount, totalPlanCountMsg, minAdmissionScoreLine, minAdmissionRank]
        result_college_list.append(row)
    return result_college_list


# 生成excel
def export_excel(result_college_list):
    # 创建一个新的工作簿
    wb = Workbook()
    # 选择默认的工作表
    ws = wb.active
    # 定义表头
    ws['A1'] = "序号"
    ws['B1'] = "地区"
    ws['C1'] = "高校名称"
    ws['D1'] = "专业组"
    ws['E1'] = "专业方向"
    ws['F1'] = "录取数"
    ws['G1'] = "录取数详情"
    ws['H1'] = "最低录取分数线"
    ws['I1'] = "最低录取排名"
    # 循环插入行数据
    for each_result_college in result_college_list:
        ws.append(each_result_college)
    # 保存工作簿到文件
    # 生成文件名
    filename = f"result-{time_str}.xlsx"
    wb.save(filename)
    print(f"导出数据到文件：{filename}")


if __name__ == '__main__':
    url = "https://www.test.com"
    headers = {
        "Accept": "application/json, text/plain, */*",
        "Accept-Encoding": "gzip, deflate, br, zstd",
        "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8",
        "Connection": "keep-alive",
        "Content-Length": "564",
        "Content-Type": "application/json",
        "Cookie": "xxxx",
        "Host": "gk-stu.bjeea.cn",
        "Nazca-Rpc-Serialize-Method": "JSON",
        "Origin": "https://www.test.com",
        "Referer": "https://www.test.com",
        "Sec-Ch-Ua": "\"Not/A)Brand\";v=\"8\", \"Chromium\";v=\"126\", \"Google Chrome\";v=\"126\"",
        "Sec-Ch-Ua-Mobile": "?0",
        "Sec-Ch-Ua-Platform": "\"Windows\"",
        "Sec-Fetch-Dest": "empty",
        "Sec-Fetch-Mode": "cors",
        "Sec-Fetch-Site": "same-origin",
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36"
    }
    # 循环次数,取决于总共多少页，对方的接口修改每页的行数不起作用，所以只能按照页数加载
    num_requests = 29
    # 循环加载数据
    concat_response_list = send_requests_and_save_responses_loop(url, "", headers, num_requests)
    # 将拼接后的字符串转换成json写入本地文件
    response_write_file(concat_response_list)
    # 响应数据格式化
    result_college_list = response_data_format(concat_response_list)
    # 导出excel
    export_excel(result_college_list)
